import openpyxl as op
from category import clean
import pickle

if __name__ == '__main__':
    ws = op.load_workbook('小纸团PPB7.11版三级目录分类.xlsx')['3.0纸团7.11终版']
    names = []
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 1).value:
            first = ws.cell(i, 1).value
        if ws.cell(i, 2).value:
            second = ws.cell(i, 2).value
        for j in range(3, ws.max_column + 1):
            if ws.cell(i, j).value:
                names.append((first, second, ws.cell(i, j).value))
    result = {}
    for name in names:
        words = name[2].split('/')
        words_cleaned = []
        for word in words:
            if len(word) == 0:
                continue
            words_cleaned.append(clean(word))
        result[name] = words_cleaned
    with open('category_separated.pkl', 'wb') as f:
        pickle.dump(result, f)
