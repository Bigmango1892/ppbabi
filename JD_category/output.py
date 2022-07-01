import pandas as pd
import openpyxl as op
import pickle

category_path = 'D:/Codes/Zhilian_Data/ppbcate/renew/分类目录.xlsx'
JD_path = 'D:/Codes/ppbabi/ppbabi/renew/jd_nona.csv'

ws = op.load_workbook(category_path)['Sheet2']
category = [ws.cell(i, j).value for i in range(3, 110) for j in range(2, 104) if ws.cell(i, j).value]

JD = pd.read_csv(JD_path)

index = [[] for _ in range(len(category))]
for i in range(len(JD)):
    for j in range(len(category)):
        if category[j] in JD.loc[i, '岗位具体名称']:
            index[j].append(i)
            break
    if i % 1000 == 0:
        print(i)

index_nona = {}
for i in range(len(index)):
    if index[i]:
        index_nona[category[i]] = index[i]

with open('JD岗位分类.data', 'wb') as f:
    pickle.dump(index_nona, f)
